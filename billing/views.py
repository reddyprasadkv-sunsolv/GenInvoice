import calendar
import logging
from collections import Counter
from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import views as auth_views
from django.core.exceptions import PermissionDenied
from django.db import IntegrityError, OperationalError, ProgrammingError, transaction
from django.http import FileResponse
from django.db.models import Count, DecimalField, F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.urls import reverse_lazy
from django.views import View
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DetailView, ListView, UpdateView

from .backup import (
    BackupError,
    RestoreError,
    RestoreValidationError,
    backup_file_path,
    cleanup_pending_restore,
    create_local_backup,
    list_local_backups,
    pending_restore_path,
    restore_local_backup,
    save_pending_restore,
)
from .forms import (
    ApplicationSettingsForm,
    BackupUploadForm,
    ClientForm,
    CompanyForm,
    DashboardFilterForm,
    DeveloperPaymentForm,
    DeveloperVendorForm,
    FirstTimeAdminSetupForm,
    HsnSacCodeForm,
    InvoiceFilterForm,
    InvoiceForm,
    InvoiceItemFormSet,
    PaymentForm,
    ProjectAssignmentForm,
    ProjectClientPaymentForm,
    ProjectForm,
    ProjectReportFilterForm,
    RecurringInvoiceTemplateForm,
    RecurringInvoiceTemplateItemFormSet,
    ReportFilterForm,
)
from .models import (
    ActivityLog,
    ApplicationSetting,
    Client,
    Company,
    CURRENCY_CHOICES,
    CURRENCY_INR,
    DeveloperPayment,
    DeveloperVendor,
    HsnSacCode,
    Invoice,
    InvoiceItem,
    Payment,
    Project,
    ProjectAssignment,
    ProjectClientPayment,
    RecurringInvoiceTemplate,
)
from .services import (
    PDFGenerationError,
    calculate_invoice_totals,
    client_fund_status,
    developer_fund_status,
    generate_draft_invoice_number,
    generate_invoice_number,
    generate_invoice_pdf,
    get_default_gst_percentage,
    invoice_title,
    project_financial_summary,
    project_gst_display_summary,
    project_report_rows,
    recalculate_assignment_payments,
    recalculate_project_client_payments,
    recalculate_invoice_payments,
    report_file_path,
    safe_invoice_pdf_filename,
    to_money,
)
from .templatetags.currency_filters import format_currency


logger = logging.getLogger(__name__)

PAGE_SIZE_CHOICES = (10, 25, 50)


class PageSizeMixin:
    paginate_by = 10

    def get_paginate_by(self, queryset):
        try:
            page_size = int(self.request.GET.get("page_size", self.paginate_by))
        except (TypeError, ValueError):
            page_size = self.paginate_by
        return page_size if page_size in PAGE_SIZE_CHOICES else self.paginate_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_size"] = self.get_paginate_by(self.object_list)
        context["page_size_choices"] = PAGE_SIZE_CHOICES
        return context


def log_activity(request, action, module, description="", object_id=None):
    try:
        ActivityLog.objects.create(
            action=action,
            module=module,
            description=(description or "")[:500],
            object_id=object_id,
            created_by=request.user if getattr(request, "user", None) and request.user.is_authenticated else None,
        )
    except (OperationalError, ProgrammingError):
        return


def _has_existing_users():
    try:
        return get_user_model().objects.exists()
    except (OperationalError, ProgrammingError):
        return True


class FirstTimeAwareLoginView(auth_views.LoginView):
    def dispatch(self, request, *args, **kwargs):
        if not _has_existing_users():
            return redirect("first_time_setup")
        return super().dispatch(request, *args, **kwargs)


@never_cache
def first_time_admin_setup(request):
    if _has_existing_users():
        messages.info(request, "First-time setup is already complete. Please log in.")
        return redirect("login")

    form = FirstTimeAdminSetupForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Admin account created. Please log in.")
        return redirect("login")

    return render(request, "registration/first_time_setup.html", {"form": form})


def superuser_required(view_func):
    @login_required
    def wrapped(request, *args, **kwargs):
        if not request.user.is_superuser:
            raise PermissionDenied("This page is restricted to administrators.")
        return view_func(request, *args, **kwargs)

    return wrapped


class RootRedirectView(View):
    def get(self, request):
        if not _has_existing_users():
            return redirect("first_time_setup")
        if request.user.is_authenticated:
            return redirect("dashboard")
        return redirect("login")


@login_required
def dashboard(request):
    filter_data = request.GET.copy()
    if not filter_data:
        filter_data["period"] = "this_month"
    filter_form = DashboardFilterForm(filter_data)
    filter_form.is_valid()
    cleaned_filters = filter_form.cleaned_data if filter_form.is_valid() else {}
    invoices = _apply_dashboard_filters(
        _final_invoice_queryset().select_related("company", "client"),
        cleaned_filters,
    )
    dashboard_projects = _dashboard_project_queryset(cleaned_filters)
    summary = _invoice_summary(invoices)
    project_summary = _project_summary(dashboard_projects)
    invoice_currency_summaries = _invoice_summary_by_currency(invoices)
    project_currency_summaries = _project_summary_by_currency(dashboard_projects, cleaned_filters, invoices)
    summary["invoice_received"] = summary["received"]
    summary["client_project_received"] = _dashboard_project_client_received_total(dashboard_projects, cleaned_filters, invoices)
    summary["total_received"] = to_money(summary["invoice_received"] + summary["client_project_received"])
    summary["received"] = summary["total_received"]
    summary["invoice_pending"] = summary["pending"]
    summary["client_project_pending"] = to_money(project_summary["total_with_gst"] - summary["client_project_received"])
    if summary["client_project_pending"] < Decimal("0.00"):
        summary["client_project_pending"] = Decimal("0.00")
    draft_invoice_count = _apply_dashboard_filters(
        Invoice.objects.filter(invoice_status=Invoice.InvoiceStatus.DRAFT, is_deleted=False),
        cleaned_filters,
    ).count()
    summary["draft_count"] = draft_invoice_count
    overdue_invoice_count = invoices.exclude(payment_status=Invoice.PaymentStatus.PAID).filter(invoice_date__lt=timezone.localdate() - timedelta(days=30)).count()
    deleted_invoice_count = _apply_dashboard_filters(
        Invoice.objects.filter(is_deleted=True),
        cleaned_filters,
    ).count()
    gst_invoice_total = _apply_gst_type_filter(invoices, "gst").aggregate(total=Sum("total_amount")).get("total") or Decimal("0.00")
    non_gst_invoice_total = _apply_gst_type_filter(invoices, "non_gst").aggregate(total=Sum("total_amount")).get("total") or Decimal("0.00")
    summary_cards = _dashboard_currency_summary_cards(invoice_currency_summaries, project_currency_summaries) + [
        {"label": "Total invoices", "value": str(summary["count"])},
        {"label": "Draft invoices", "value": str(draft_invoice_count)},
        {"label": "Deleted invoices", "value": str(deleted_invoice_count)},
        {"label": "Paid invoices", "value": str(summary["paid_count"])},
        {"label": "Pending invoices", "value": str(summary["pending_count"])},
        {"label": "Partially paid invoices", "value": str(summary["partial_count"])},
        {"label": "Overdue Payments", "value": str(overdue_invoice_count)},
        {"label": "GST collected / payable", "value": _money(summary["gst"])},
        {"label": "GST invoice total", "value": _money(gst_invoice_total)},
        {"label": "Non-GST invoice total", "value": _money(non_gst_invoice_total)},
    ]
    recent_invoices = invoices.order_by("-invoice_date", "-created_at")[:8]
    project_summary_cards = _project_summary_card_list(project_summary)
    analytics_cards = _dashboard_analytics_cards(summary, project_summary, invoice_currency_summaries, project_currency_summaries)
    dashboard_chart_data = _dashboard_chart_data(invoices, dashboard_projects, project_summary, summary)
    return render(
        request,
        "billing/dashboard.html",
        {
            "filter_form": filter_form,
            "summary_cards": summary_cards,
            "received_amount": summary["received"],
            "invoice_received_amount": summary["invoice_received"],
            "client_project_received_amount": summary["client_project_received"],
            "total_received_amount": summary["total_received"],
            "pending_amount": summary["pending"],
            "invoice_pending_amount": summary["invoice_pending"],
            "client_project_pending_amount": summary["client_project_pending"],
            "total_invoice_amount": summary["raised"],
            "gst_amount": summary["gst"],
            "invoice_payment_status_data": dashboard_chart_data["invoicePaymentStatus"],
            "project_summary_cards": project_summary_cards,
            "analytics_cards": analytics_cards,
            "dashboard_chart_data": dashboard_chart_data,
            "invoice_currency_summaries": invoice_currency_summaries,
            "project_currency_summaries": project_currency_summaries,
            "recent_invoices": recent_invoices,
            "backup_reminder": _backup_reminder_context(request),
            "recent_activities": ActivityLog.objects.select_related("created_by").exclude(action__in=["deleted", "restored"])[:5],
            "upcoming_recurring_templates": RecurringInvoiceTemplate.objects.select_related("company", "client", "project").filter(
                is_active=True,
                next_invoice_date__lte=timezone.localdate() + timedelta(days=30),
            )[:5],
        },
    )


def _money(value, currency=CURRENCY_INR):
    return format_currency(value, currency)


class CompanyListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = Company
    template_name = "billing/company_list.html"
    context_object_name = "companies"


class CompanyDetailView(LoginRequiredMixin, DetailView):
    model = Company
    template_name = "billing/company_detail.html"
    context_object_name = "company"


class CompanyCreateView(LoginRequiredMixin, CreateView):
    model = Company
    form_class = CompanyForm
    template_name = "billing/company_form.html"
    success_url = reverse_lazy("company_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        log_activity(self.request, "created", "Company", f"Company created: {self.object.company_name}", self.object.pk)
        messages.success(self.request, "Company details saved.")
        return response


class CompanyUpdateView(LoginRequiredMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = "billing/company_form.html"

    def get_success_url(self):
        return reverse_lazy("company_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        log_activity(self.request, "updated", "Company", f"Company updated: {self.object.company_name}", self.object.pk)
        messages.success(self.request, "Company details updated.")
        return response


class ClientListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = Client
    template_name = "billing/client_list.html"
    context_object_name = "clients"

    def get_queryset(self):
        self.status_filter = self.request.GET.get("status", "active")
        queryset = Client.objects.all()
        if self.status_filter == "deleted":
            queryset = queryset.filter(Q(is_deleted=True) | Q(client_status=Client.ClientStatus.DELETED))
        elif self.status_filter == "draft":
            queryset = queryset.filter(is_deleted=False, client_status=Client.ClientStatus.DRAFT)
        elif self.status_filter == "all":
            queryset = queryset
        else:
            queryset = queryset.filter(is_deleted=False, client_status=Client.ClientStatus.ACTIVE)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["status_filter"] = getattr(self, "status_filter", "active")
        return context


class ClientDetailView(LoginRequiredMixin, DetailView):
    model = Client
    template_name = "billing/client_detail.html"
    context_object_name = "client_record"


@login_required
def client_ledger(request, pk):
    client = get_object_or_404(Client, pk=pk)
    invoices = client.invoices.select_related("company", "project").prefetch_related("payments").filter(is_deleted=False)
    final_invoices = invoices.filter(invoice_status=Invoice.InvoiceStatus.FINAL)
    payments = Payment.objects.select_related("invoice", "invoice__company").filter(
        invoice__client=client,
        invoice__is_deleted=False,
        invoice__invoice_status=Invoice.InvoiceStatus.FINAL,
    )
    summary = _invoice_summary(final_invoices)
    currency_summaries = _report_currency_summaries(final_invoices, payments, "payment_received")
    gst_invoice_total = _apply_gst_type_filter(final_invoices, "gst").aggregate(total=Sum("total_amount")).get("total") or Decimal("0.00")
    non_gst_invoice_total = _apply_gst_type_filter(final_invoices, "non_gst").aggregate(total=Sum("total_amount")).get("total") or Decimal("0.00")
    return render(
        request,
        "billing/client_ledger.html",
        {
            "client_record": client,
            "summary": summary,
            "currency_summaries": currency_summaries,
            "gst_invoice_total": gst_invoice_total,
            "non_gst_invoice_total": non_gst_invoice_total,
            "invoices": final_invoices,
            "draft_invoices": invoices.filter(invoice_status=Invoice.InvoiceStatus.DRAFT),
            "payments": payments,
            "projects": client.projects.prefetch_related("client_payments", "assignments__developer_payments"),
            "pending_receivables": final_invoices.filter(pending_amount__gt=0),
        },
    )


class ClientCreateView(LoginRequiredMixin, CreateView):
    model = Client
    form_class = ClientForm
    template_name = "billing/client_form.html"
    success_url = reverse_lazy("client_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["is_draft"] = self.request.method == "POST" and self.request.POST.get("client_action") == "draft"
        return kwargs

    def form_valid(self, form):
        self.object = form.save(commit=False)
        is_draft = self.request.POST.get("client_action") == "draft"
        self.object.client_status = Client.ClientStatus.DRAFT if is_draft else Client.ClientStatus.ACTIVE
        self.object.save()
        log_activity(self.request, "created", "Client", f"Client created: {self.object.client_name}", self.object.pk)
        messages.success(self.request, "Draft client saved." if is_draft else "Client details saved.")
        return redirect(self.get_success_url())


class ClientUpdateView(LoginRequiredMixin, UpdateView):
    model = Client
    form_class = ClientForm
    template_name = "billing/client_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["is_draft"] = self.request.method == "POST" and self.request.POST.get("client_action") == "draft"
        return kwargs

    def get_success_url(self):
        return reverse_lazy("client_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        self.object = form.save(commit=False)
        is_draft = self.request.POST.get("client_action") == "draft"
        if is_draft:
            self.object.client_status = Client.ClientStatus.DRAFT
        elif not self.object.is_deleted:
            self.object.client_status = Client.ClientStatus.ACTIVE
        self.object.save()
        log_activity(self.request, "updated", "Client", f"Client updated: {self.object.client_name}", self.object.pk)
        messages.success(self.request, "Draft client saved." if is_draft else "Client details updated.")
        return redirect(self.get_success_url())


class HsnSacCodeListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = HsnSacCode
    template_name = "billing/hsn_sac_code_list.html"
    context_object_name = "codes"


class HsnSacCodeCreateView(LoginRequiredMixin, CreateView):
    model = HsnSacCode
    form_class = HsnSacCodeForm
    template_name = "billing/hsn_sac_code_form.html"
    success_url = reverse_lazy("hsn_sac_code_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        log_activity(self.request, "created", "HSN/SAC Code", f"HSN/SAC code created: {self.object.code}", self.object.pk)
        messages.success(self.request, "HSN/SAC Code saved.")
        return response


class HsnSacCodeUpdateView(LoginRequiredMixin, UpdateView):
    model = HsnSacCode
    form_class = HsnSacCodeForm
    template_name = "billing/hsn_sac_code_form.html"
    success_url = reverse_lazy("hsn_sac_code_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        log_activity(self.request, "updated", "HSN/SAC Code", f"HSN/SAC code updated: {self.object.code}", self.object.pk)
        messages.success(self.request, "HSN/SAC Code updated.")
        return response


@login_required
def hsn_sac_code_delete(request, pk):
    code = get_object_or_404(HsnSacCode, pk=pk)
    if request.method == "POST":
        label = code.code
        code.delete()
        log_activity(request, "deleted", "HSN/SAC Code", f"HSN/SAC code deleted: {label}", pk)
        messages.success(request, f"HSN/SAC Code deleted: {label}")
        return redirect("hsn_sac_code_list")
    return render(
        request,
        "billing/confirm_delete.html",
        {
            "object_type": "HSN/SAC code",
            "object_name": code.code,
            "cancel_url": reverse_lazy("hsn_sac_code_list"),
            "linked_records": InvoiceItem.objects.filter(hsn_sac_code=code).count(),
            "warning_message": "Existing invoice items using this code will keep working and show no HSN/SAC code after deletion.",
            "confirm_label": "Delete",
            "confirmation_text": "Are you sure you want to delete this HSN/SAC code?",
        },
    )


@login_required
def client_soft_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":
        client.is_deleted = True
        client.client_status = Client.ClientStatus.DELETED
        client.deleted_at = timezone.now()
        client.deleted_by = request.user
        client.save(update_fields=["client_status", "is_deleted", "deleted_at", "deleted_by", "updated_at"])
        log_activity(request, "deleted", "Client", f"Client soft deleted: {client.client_name}", client.pk)
        messages.warning(request, f"Client soft deleted: {client.client_name}")
        return redirect("client_list")
    return render(
        request,
        "billing/confirm_delete.html",
        {
            "object_type": "client",
            "object_name": client.client_name,
            "cancel_url": reverse_lazy("client_list"),
            "linked_records": client.invoices.count() + client.projects.count(),
            "warning_message": "This will hide the client from new invoice and project dropdowns. Historical records will remain available.",
            "confirm_label": "Soft Delete",
            "confirmation_text": "Are you sure you want to soft delete this client?",
        },
    )


@login_required
@require_POST
def client_restore(request, pk):
    client = get_object_or_404(Client, pk=pk)
    client.is_deleted = False
    client.client_status = Client.ClientStatus.ACTIVE
    client.deleted_at = None
    client.deleted_by = None
    client.save(update_fields=["client_status", "is_deleted", "deleted_at", "deleted_by", "updated_at"])
    log_activity(request, "restored", "Client", f"Client restored: {client.client_name}", client.pk)
    messages.success(request, f"Client restored: {client.client_name}")
    return redirect("client_list")


@login_required
def placeholder(request, section):
    return render(request, "billing/placeholder.html", {"section": section})


@login_required
def global_search(request):
    query = request.GET.get("q", "").strip()
    sections = []
    if query:
        invoice_queryset = (
            Invoice.objects.select_related("company", "client", "project")
            .filter(is_deleted=False)
            .filter(
                Q(invoice_number__icontains=query)
                | Q(subject__icontains=query)
                | Q(company__company_name__icontains=query)
                | Q(client__client_name__icontains=query)
                | Q(project__project_name__icontains=query)
                | Q(items__hsn_sac_code__code__icontains=query)
            )
            .distinct()
        )
        client_queryset = Client.objects.filter(is_deleted=False).filter(
            Q(client_name__icontains=query) | Q(city__icontains=query) | Q(gstin__icontains=query)
        )
        company_queryset = Company.objects.filter(
            Q(company_name__icontains=query) | Q(city__icontains=query) | Q(gstin__icontains=query)
        )
        project_queryset = Project.objects.select_related("client").filter(
            Q(project_name__icontains=query)
            | Q(project_id__icontains=query)
            | Q(client__client_name__icontains=query)
        )
        vendor_queryset = DeveloperVendor.objects.filter(
            Q(name__icontains=query)
            | Q(contact_person__icontains=query)
            | Q(email__icontains=query)
            | Q(gstin__icontains=query)
        )
        hsn_queryset = HsnSacCode.objects.filter(Q(code__icontains=query) | Q(description__icontains=query))
        sections = [
            _search_section("Invoices", "invoices", invoice_queryset, "invoice_detail", "invoice_edit"),
            _search_section("Clients", "clients", client_queryset, "client_detail", "client_edit"),
            _search_section("Companies", "companies", company_queryset, "company_detail", "company_edit"),
            _search_section("Projects", "projects", project_queryset, "project_detail", "project_edit"),
            _search_section("Developers/Vendors", "vendors", vendor_queryset, "developer_vendor_detail", "developer_vendor_edit"),
            _search_section("HSN/SAC Codes", "hsn", hsn_queryset, None, "hsn_sac_code_edit"),
        ]
    return render(request, "billing/search_results.html", {"query": query, "sections": sections})


def _search_section(title, key, queryset, view_name, edit_name):
    total = queryset.count()
    return {
        "title": title,
        "key": key,
        "items": queryset[:10],
        "total": total,
        "view_name": view_name,
        "edit_name": edit_name,
    }


class ActivityLogListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = ActivityLog
    template_name = "billing/activity_log.html"
    context_object_name = "activities"

    def get_queryset(self):
        return ActivityLog.objects.select_related("created_by")


class InvoiceListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = Invoice
    template_name = "billing/invoice_list.html"
    context_object_name = "invoices"
    paginate_by = 10

    def get_queryset(self):
        self.filter_form = InvoiceFilterForm(self.request.GET or None)
        queryset = Invoice.objects.select_related("company", "client", "project")
        if self.filter_form.is_valid():
            queryset = _apply_invoice_filters(queryset, self.filter_form.cleaned_data)
        else:
            queryset = queryset.filter(is_deleted=False)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = getattr(self, "filter_form", InvoiceFilterForm())
        return context


class InvoiceDetailView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = "billing/invoice_detail.html"
    context_object_name = "invoice"

    def get_queryset(self):
        return Invoice.objects.select_related("company", "client", "project").prefetch_related("items__hsn_sac_code", "payments")


class InvoicePreviewView(InvoiceDetailView):
    template_name = "billing/invoice_preview.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["invoice_title"] = invoice_title(self.object)
        return context


@login_required
def invoice_pdf_download(request, pk):
    invoice = get_object_or_404(Invoice.objects.select_related("company", "client", "project").prefetch_related("items__hsn_sac_code"), pk=pk)
    if invoice.is_deleted or invoice.invoice_status != Invoice.InvoiceStatus.FINAL:
        messages.error(request, "Only final active invoices can be downloaded as PDF.")
        return redirect("invoice_detail", pk=invoice.pk)
    try:
        pdf_path = generate_invoice_pdf(invoice, request=request)
    except PDFGenerationError as exc:
        logger.exception("Invoice PDF generation failed for invoice %s", invoice.pk)
        messages.error(request, f"PDF generation failed: {exc}")
        return redirect("invoice_detail", pk=invoice.pk)

    response = FileResponse(open(pdf_path, "rb"), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{safe_invoice_pdf_filename(invoice.invoice_number)}"'
    return response


@login_required
def invoice_create(request):
    invoice = Invoice(invoice_date=timezone.localdate())
    if request.method == "POST":
        save_as_draft = request.POST.get("invoice_action") == "draft"
        form = InvoiceForm(request.POST, instance=invoice, is_draft=save_as_draft)
        formset = InvoiceItemFormSet(request.POST, prefix="items", require_items=not save_as_draft)
        if form.is_valid() and formset.is_valid():
            invoice = form.save(commit=False)
            invoice.invoice_number = (
                generate_draft_invoice_number(invoice.invoice_date)
                if save_as_draft
                else generate_invoice_number(invoice.company, invoice.client, invoice.invoice_date)
            )
            invoice.received_amount = to_money(0)
            invoice.payment_status = Invoice.PaymentStatus.PENDING
            if _save_invoice_with_items(request, invoice, formset, created=True, save_as_draft=save_as_draft):
                log_activity(
                    request,
                    "draft created" if save_as_draft else "created",
                    "Invoice",
                    f"Invoice {'draft ' if save_as_draft else ''}created: {invoice.invoice_number}",
                    invoice.pk,
                )
                messages.success(request, "Draft invoice saved." if save_as_draft else "Invoice created.")
                return redirect("invoice_detail", pk=invoice.pk)
    else:
        form = InvoiceForm(instance=invoice)
        formset = InvoiceItemFormSet(prefix="items")

    return render(
        request,
        "billing/invoice_form.html",
        _invoice_form_context(form, formset, invoice, is_edit=False),
    )


@login_required
def invoice_edit(request, pk):
    invoice = get_object_or_404(Invoice.objects.select_related("company", "client").prefetch_related("items__hsn_sac_code"), pk=pk)
    if request.method == "POST":
        save_as_draft = request.POST.get("invoice_action") == "draft" and invoice.invoice_status == Invoice.InvoiceStatus.DRAFT
        was_draft = invoice.invoice_status == Invoice.InvoiceStatus.DRAFT
        form = InvoiceForm(request.POST, instance=invoice, is_draft=save_as_draft)
        formset = InvoiceItemFormSet(request.POST, prefix="items", require_items=not save_as_draft)
        if form.is_valid() and formset.is_valid():
            invoice = form.save(commit=False)
            if was_draft and not save_as_draft:
                invoice.invoice_number = generate_invoice_number(invoice.company, invoice.client, invoice.invoice_date)
            if _save_invoice_with_items(
                request,
                invoice,
                formset,
                created=False,
                save_as_draft=save_as_draft,
                finalizing_draft=was_draft and not save_as_draft,
            ):
                log_activity(
                    request,
                    "finalized" if was_draft and not save_as_draft else "updated",
                    "Invoice",
                    f"Invoice {'finalized' if was_draft and not save_as_draft else 'updated'}: {invoice.invoice_number}",
                    invoice.pk,
                )
                messages.success(request, "Draft invoice saved." if save_as_draft else "Invoice updated.")
                return redirect("invoice_detail", pk=invoice.pk)
    else:
        form = InvoiceForm(instance=invoice, is_draft=invoice.invoice_status == Invoice.InvoiceStatus.DRAFT)
        formset = InvoiceItemFormSet(prefix="items", initial=_invoice_item_initial(invoice), require_items=invoice.invoice_status != Invoice.InvoiceStatus.DRAFT)

    return render(
        request,
        "billing/invoice_form.html",
        _invoice_form_context(form, formset, invoice, is_edit=True),
    )


def _invoice_form_context(form, formset, invoice, is_edit):
    return {
        "form": form,
        "formset": formset,
        "invoice": invoice,
        "is_edit": is_edit,
        "company_gstin_map": {str(company.pk): bool(company.gstin) for company in Company.objects.all()},
        "client_gst_preference_map": {str(client.pk): bool(client.requires_gst_invoice) for client in Client.objects.all()},
        "project_client_map": {str(project.pk): str(project.client_id) for project in Project.objects.all()},
        "invoice_settings": {
            "defaultGstPercentage": str(get_default_gst_percentage()),
            "isEdit": is_edit,
        },
    }

def _invoice_item_initial(invoice):
    return [
        {
            "description": item.description,
            "hsn_sac_code": item.hsn_sac_code,
            "item_price": item.item_price,
            "quantity": item.quantity,
        }
        for item in invoice.items.all()
    ]


def _formset_item_rows(formset):
    rows = []
    for form in formset.forms:
        cleaned_data = getattr(form, "cleaned_data", {})
        if cleaned_data.get("DELETE"):
            continue
        has_values = any(
            cleaned_data.get(field_name) not in (None, "")
            for field_name in ("description", "item_price", "quantity")
        )
        if not has_values:
            continue
        rows.append(
            {
                "description": cleaned_data["description"],
                "hsn_sac_code": cleaned_data.get("hsn_sac_code"),
                "item_price": cleaned_data["item_price"],
                "quantity": cleaned_data["quantity"],
            }
        )
    return rows


def _save_invoice_with_items(request, invoice, formset, created, save_as_draft=False, finalizing_draft=False):
    invoice.apply_gst = bool(invoice.apply_gst and invoice.company and invoice.company.gstin)
    calculated = calculate_invoice_totals(_formset_item_rows(formset), invoice.company, apply_gst=invoice.apply_gst, currency=invoice.currency)
    invoice.subtotal = calculated["subtotal"]
    invoice.gst_percentage = calculated["gst_percentage"]
    invoice.gst_amount = calculated["gst_amount"]
    invoice.total_amount = calculated["total_amount"]
    invoice.amount_in_words = calculated["amount_in_words"]
    invoice.invoice_status = Invoice.InvoiceStatus.DRAFT if save_as_draft else Invoice.InvoiceStatus.FINAL
    if created or finalizing_draft:
        invoice.received_amount = to_money(0)
    invoice.pending_amount = to_money(invoice.total_amount - invoice.received_amount)
    invoice.payment_status = _payment_status(invoice)
    if save_as_draft:
        invoice.received_amount = to_money(0)
        invoice.pending_amount = to_money(invoice.total_amount)
        invoice.payment_status = Invoice.PaymentStatus.PENDING

    try:
        with transaction.atomic():
            invoice.save()
            InvoiceItem.objects.filter(invoice=invoice).delete()
            InvoiceItem.objects.bulk_create(
                [
                    InvoiceItem(
                        invoice=invoice,
                        serial_number=item["serial_number"],
                        description=item["description"],
                        hsn_sac_code=item.get("hsn_sac_code"),
                        item_price=item["item_price"],
                        quantity=item["quantity"],
                        total=item["total"],
                    )
                    for item in calculated["items"]
                ]
            )
            if finalizing_draft:
                Payment.objects.filter(invoice=invoice).delete()
            recalculate_invoice_payments(invoice)
    except IntegrityError:
        messages.error(request, "Invoice number already exists. Please try again.")
        return False
    return True


def _payment_status(invoice):
    if invoice.received_amount >= invoice.total_amount and invoice.total_amount > 0:
        return Invoice.PaymentStatus.PAID
    if invoice.received_amount > 0:
        return Invoice.PaymentStatus.PARTIALLY_PAID
    return Invoice.PaymentStatus.PENDING


@login_required
def invoice_add_payment(request, pk):
    invoice = get_object_or_404(Invoice.objects.select_related("company", "client"), pk=pk)
    if invoice.is_deleted or invoice.invoice_status != Invoice.InvoiceStatus.FINAL:
        messages.error(request, "Payments can be recorded only for final active invoices.")
        return redirect("invoice_detail", pk=invoice.pk)
    if request.method == "POST":
        form = PaymentForm(request.POST, invoice=invoice)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.invoice = invoice
            with transaction.atomic():
                payment.save()
                recalculate_invoice_payments(invoice)
            log_activity(request, "payment added", "Invoice", f"Payment recorded for invoice: {invoice.invoice_number}", invoice.pk)
            messages.success(request, "Payment recorded.")
            return redirect("invoice_detail", pk=invoice.pk)
    else:
        form = PaymentForm(invoice=invoice, initial={"payment_date": timezone.localdate()})
    return render(request, "billing/payment_form.html", {"form": form, "invoice": invoice})


@login_required
def invoice_payment_history(request, pk):
    invoice = get_object_or_404(
        Invoice.objects.select_related("company", "client").prefetch_related("payments"),
        pk=pk,
    )
    return render(request, "billing/payment_history.html", {"invoice": invoice})


@login_required
def invoice_soft_delete(request, pk):
    invoice = get_object_or_404(Invoice.objects.select_related("company", "client"), pk=pk)
    if request.method == "POST":
        invoice.is_deleted = True
        invoice.invoice_status = Invoice.InvoiceStatus.DELETED
        invoice.deleted_at = timezone.now()
        invoice.deleted_by = request.user
        invoice.save(update_fields=["is_deleted", "invoice_status", "deleted_at", "deleted_by", "updated_at"])
        log_activity(request, "deleted", "Invoice", f"Invoice soft deleted: {invoice.invoice_number}", invoice.pk)
        messages.warning(request, f"Invoice soft deleted: {invoice.invoice_number}")
        return redirect("invoice_list")
    return render(
        request,
        "billing/confirm_delete.html",
        {
            "object_type": "invoice",
            "object_name": invoice.invoice_number,
            "cancel_url": reverse_lazy("invoice_list"),
            "linked_records": invoice.payments.count(),
            "warning_message": "This will hide the invoice from dashboard totals and default reports. Historical data and PDF files remain available.",
            "confirm_label": "Soft Delete",
            "confirmation_text": "Are you sure you want to soft delete this invoice?",
        },
    )


@login_required
@require_POST
def invoice_restore(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice.is_deleted = False
    invoice.deleted_at = None
    invoice.deleted_by = None
    if invoice.invoice_status == Invoice.InvoiceStatus.DELETED:
        invoice.invoice_status = (
            Invoice.InvoiceStatus.DRAFT if invoice.invoice_number.startswith("DRAFT-") else Invoice.InvoiceStatus.FINAL
        )
    invoice.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "invoice_status", "updated_at"])
    log_activity(request, "restored", "Invoice", f"Invoice restored: {invoice.invoice_number}", invoice.pk)
    messages.success(request, f"Invoice restored: {invoice.invoice_number}")
    return redirect("invoice_list")


@login_required
@require_POST
def invoice_clone(request, pk):
    source = get_object_or_404(
        Invoice.objects.select_related("company", "client", "project").prefetch_related("items__hsn_sac_code"),
        pk=pk,
    )
    invoice_date = timezone.localdate()
    item_rows = [
        {
            "description": item.description,
            "hsn_sac_code": item.hsn_sac_code,
            "item_price": item.item_price,
            "quantity": item.quantity,
        }
        for item in source.items.all()
    ]
    calculated = calculate_invoice_totals(item_rows, source.company, apply_gst=source.apply_gst, currency=source.currency)
    clone = Invoice(
        invoice_number=generate_draft_invoice_number(invoice_date),
        company=source.company,
        client=source.client,
        project=source.project,
        invoice_date=invoice_date,
        currency=source.currency,
        subject=source.subject,
        apply_gst=bool(source.apply_gst and source.company.gstin),
        subtotal=calculated["subtotal"],
        gst_percentage=calculated["gst_percentage"],
        gst_amount=calculated["gst_amount"],
        total_amount=calculated["total_amount"],
        amount_in_words=calculated["amount_in_words"],
        terms_and_conditions=source.terms_and_conditions,
        declaration=source.declaration,
        payment_status=Invoice.PaymentStatus.PENDING,
        invoice_status=Invoice.InvoiceStatus.DRAFT,
        received_amount=to_money(0),
        pending_amount=calculated["total_amount"],
    )
    with transaction.atomic():
        clone.save()
        InvoiceItem.objects.bulk_create(
            [
                InvoiceItem(
                    invoice=clone,
                    serial_number=item["serial_number"],
                    description=item["description"],
                    hsn_sac_code=item.get("hsn_sac_code"),
                    item_price=item["item_price"],
                    quantity=item["quantity"],
                    total=item["total"],
                )
                for item in calculated["items"]
            ]
        )
    log_activity(request, "cloned", "Invoice", f"Invoice cloned from {source.invoice_number} to {clone.invoice_number}", clone.pk)
    messages.success(request, f"Draft invoice created from {source.invoice_number}.")
    return redirect("invoice_edit", pk=clone.pk)


class ProjectListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = Project
    template_name = "billing/project_list.html"
    context_object_name = "projects"

    def get_queryset(self):
        return Project.objects.select_related("client").prefetch_related(
            "assignments__developer_vendor",
            "assignments__developer_payments",
            "client_payments",
            "invoices",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["project_rows"] = project_report_rows(context["projects"])
        return context


class ProjectDetailView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = "billing/project_detail.html"
    context_object_name = "project"

    def get_queryset(self):
        return Project.objects.select_related("client").prefetch_related(
            "client_payments",
            "assignments__developer_vendor",
            "assignments__developer_payments",
            "invoices__company",
            "invoices__client",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        summary = project_financial_summary(self.object)
        gst_summary = project_gst_display_summary(self.object)
        context["summary"] = summary
        context["project_gst_summary"] = gst_summary
        context["project_base_amount"] = gst_summary["base_amount"]
        context["project_gst_amount"] = gst_summary["gst_amount"]
        context["project_total_with_gst"] = gst_summary["total_with_gst"]
        context["project_client_pending_amount"] = gst_summary["client_pending"]
        context["client_fund_status"] = client_fund_status(
            summary["client_received"],
            gst_summary["client_pending"],
            gst_summary["total_with_gst"],
        )
        context["developer_fund_status"] = developer_fund_status(
            summary["developer_paid"],
            summary["developer_pending"],
            summary["developer_final_cost"],
        )
        return context


@login_required
def project_ledger(request, pk):
    project = get_object_or_404(
        Project.objects.select_related("client").prefetch_related(
            "client_payments",
            "assignments__developer_vendor",
            "assignments__developer_payments",
            "invoices__company",
            "invoices__payments",
        ),
        pk=pk,
    )
    summary = project_financial_summary(project)
    invoices = project.invoices.filter(is_deleted=False).select_related("company", "client")
    return render(
        request,
        "billing/project_ledger.html",
        {
            "project": project,
            "summary": summary,
            "client_fund_status": client_fund_status(
                project.client_total_amount_received,
                project.client_pending_amount,
                summary["total_with_gst"],
            ),
            "developer_fund_status": developer_fund_status(
                summary["developer_paid"],
                summary["developer_pending"],
                summary["developer_final_cost"],
            ),
            "invoices": invoices,
        },
    )


class ProjectCreateView(LoginRequiredMixin, CreateView):
    model = Project
    form_class = ProjectForm
    template_name = "billing/project_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["is_draft"] = self.request.method == "POST" and self.request.POST.get("project_action") == "draft"
        return kwargs

    def get_success_url(self):
        return reverse_lazy("project_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        recalculate_project_client_payments(self.object)
        log_activity(self.request, "created", "Project", f"Project created: {self.object.project_name}", self.object.pk)
        messages.success(self.request, "Draft project saved." if self.object.project_status == Project.ProjectStatus.DRAFT else "Project created.")
        return response


class ProjectUpdateView(LoginRequiredMixin, UpdateView):
    model = Project
    form_class = ProjectForm
    template_name = "billing/project_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["is_draft"] = self.request.method == "POST" and self.request.POST.get("project_action") == "draft"
        return kwargs

    def get_success_url(self):
        return reverse_lazy("project_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        recalculate_project_client_payments(self.object)
        log_activity(self.request, "updated", "Project", f"Project updated: {self.object.project_name}", self.object.pk)
        messages.success(self.request, "Draft project saved." if self.object.project_status == Project.ProjectStatus.DRAFT else "Project updated.")
        return response


@login_required
def project_add_client_payment(request, pk):
    project = get_object_or_404(Project.objects.select_related("client"), pk=pk)
    if request.method == "POST":
        form = ProjectClientPaymentForm(request.POST, project=project)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.project = project
            with transaction.atomic():
                payment.save()
                recalculate_project_client_payments(project)
            log_activity(request, "payment added", "Project", f"Client payment recorded for project: {project.project_name}", project.pk)
            messages.success(request, "Client project payment recorded.")
            return redirect("project_detail", pk=project.pk)
    else:
        form = ProjectClientPaymentForm(project=project, initial={"payment_date": timezone.localdate()})
    return render(request, "billing/project_client_payment_form.html", {"form": form, "project": project, "is_edit": False})


@login_required
def project_edit_client_payment(request, project_pk, payment_pk):
    payment = get_object_or_404(
        ProjectClientPayment.objects.select_related("project", "project__client"),
        pk=payment_pk,
        project_id=project_pk,
    )
    project = payment.project
    if request.method == "POST":
        form = ProjectClientPaymentForm(request.POST, project=project, instance=payment)
        if form.is_valid():
            with transaction.atomic():
                form.save()
                recalculate_project_client_payments(project)
            log_activity(request, "payment updated", "Project", f"Client payment updated for project: {project.project_name}", project.pk)
            messages.success(request, "Client project payment updated.")
            return redirect("project_detail", pk=project.pk)
    else:
        form = ProjectClientPaymentForm(project=project, instance=payment)
    return render(
        request,
        "billing/project_client_payment_form.html",
        {"form": form, "project": project, "payment": payment, "is_edit": True},
    )


@login_required
def project_assign_developer(request, pk):
    project = get_object_or_404(Project.objects.select_related("client"), pk=pk)
    if request.method == "POST":
        form = ProjectAssignmentForm(request.POST)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.project = project
            assignment.developer_cost_estimate = assignment.developer_cost_estimate or Decimal("0.00")
            assignment.developer_final_project_cost = assignment.developer_final_project_cost or Decimal("0.00")
            assignment.advance_amount_sent = assignment.advance_amount_sent or Decimal("0.00")
            assignment.next_advance_amount_to_send = assignment.next_advance_amount_to_send or Decimal("0.00")
            with transaction.atomic():
                assignment.save()
                recalculate_assignment_payments(assignment)
            log_activity(request, "developer assigned", "Project", f"Developer/vendor assigned to project: {project.project_name}", project.pk)
            messages.success(request, "Developer/vendor assigned to project.")
            return redirect("project_detail", pk=project.pk)
    else:
        form = ProjectAssignmentForm()
    return render(request, "billing/project_assignment_form.html", {"form": form, "project": project})


@login_required
def assignment_add_developer_payment(request, pk):
    assignment = get_object_or_404(
        ProjectAssignment.objects.select_related("project", "project__client", "developer_vendor"),
        pk=pk,
    )
    if request.method == "POST":
        form = DeveloperPaymentForm(request.POST, assignment=assignment)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.project_assignment = assignment
            with transaction.atomic():
                payment.save()
                recalculate_assignment_payments(assignment)
            log_activity(request, "payment added", "Project", f"Developer/vendor payment recorded for project: {assignment.project.project_name}", assignment.project.pk)
            messages.success(request, "Developer/vendor payment recorded.")
            return redirect("project_detail", pk=assignment.project.pk)
    else:
        form = DeveloperPaymentForm(assignment=assignment, initial={"payment_date": timezone.localdate()})
    return render(request, "billing/developer_payment_form.html", {"form": form, "assignment": assignment})


@login_required
@require_POST
def project_assignment_delete(request, project_id, assignment_id):
    project = get_object_or_404(Project, pk=project_id)
    assignment = get_object_or_404(
        ProjectAssignment.objects.select_related("developer_vendor", "project").prefetch_related("developer_payments"),
        pk=assignment_id,
        project=project,
    )
    if assignment.developer_payments.exists():
        messages.error(
            request,
            "This assignment has payment records and cannot be deleted. Please remove related payments first or mark the assignment as inactive/cancelled.",
        )
        return redirect("project_detail", pk=project.pk)

    vendor_name = assignment.developer_vendor.name
    assignment.delete()
    log_activity(request, "deleted", "Project", f"Developer/vendor assignment deleted from project: {project.project_name} - {vendor_name}", project.pk)
    messages.success(request, "Developer/Vendor assignment deleted successfully.")
    return redirect("project_detail", pk=project.pk)


@login_required
def project_invoices(request, pk):
    project = get_object_or_404(Project.objects.select_related("client"), pk=pk)
    invoices = project.invoices.filter(is_deleted=False).select_related("company", "client", "project").prefetch_related("payments")
    return render(request, "billing/project_invoices.html", {"project": project, "invoices": invoices})


@login_required
def project_delete(request, pk):
    project = get_object_or_404(Project.objects.select_related("client"), pk=pk)
    linked_records = _project_linked_record_count(project)
    if request.method == "POST":
        if linked_records:
            project.project_status = Project.ProjectStatus.CANCELLED
            project.save(update_fields=["project_status", "completion_percentage", "updated_at"])
            log_activity(request, "deleted", "Project", f"Project marked cancelled: {project.project_name}", project.pk)
            messages.warning(
                request,
                "This project has linked invoices/payments/assignments and cannot be permanently deleted. It was marked as Cancelled.",
            )
        else:
            project_name = project.project_name
            project.delete()
            log_activity(request, "deleted", "Project", f"Project deleted: {project_name}", pk)
            messages.success(request, f"Project deleted: {project_name}")
        return redirect("project_list")
    return render(
        request,
        "billing/confirm_delete.html",
        {
            "object_type": "project",
            "object_name": project.project_name,
            "cancel_url": reverse_lazy("project_list"),
            "linked_records": linked_records,
            "warning_message": (
                "This project has linked invoices/payments/assignments and cannot be permanently deleted. "
                "Confirming will mark it as Cancelled."
            )
            if linked_records
            else "",
            "confirm_label": "Mark Cancelled" if linked_records else "Confirm Delete",
            "confirmation_text": "Are you sure you want to delete this project? This action cannot be undone if no records are linked.",
        },
    )


class DeveloperVendorListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = DeveloperVendor
    template_name = "billing/developer_vendor_list.html"
    context_object_name = "vendors"

    def get_queryset(self):
        return DeveloperVendor.objects.annotate(assignment_count=Count("assignments", distinct=True)).order_by("name")


class DeveloperVendorDetailView(LoginRequiredMixin, DetailView):
    model = DeveloperVendor
    template_name = "billing/developer_vendor_detail.html"
    context_object_name = "vendor"

    def get_queryset(self):
        return DeveloperVendor.objects.prefetch_related("assignments__project")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rows = []
        for assignment in self.object.assignments.select_related("project"):
            rows.append(
                {
                    "assignment": assignment,
                    "fund_status": developer_fund_status(
                        assignment.total_amount_paid_to_developer,
                        assignment.pending_amount_to_developer,
                        assignment.developer_final_project_cost or assignment.developer_cost_estimate,
                    ),
                }
            )
        context["assignment_rows"] = rows
        return context


class DeveloperVendorCreateView(LoginRequiredMixin, CreateView):
    model = DeveloperVendor
    form_class = DeveloperVendorForm
    template_name = "billing/developer_vendor_form.html"
    success_url = reverse_lazy("developer_vendor_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        log_activity(self.request, "created", "Developer/Vendor", f"Developer/vendor created: {self.object.name}", self.object.pk)
        messages.success(self.request, "Developer/vendor saved.")
        return response


class DeveloperVendorUpdateView(LoginRequiredMixin, UpdateView):
    model = DeveloperVendor
    form_class = DeveloperVendorForm
    template_name = "billing/developer_vendor_form.html"

    def get_success_url(self):
        return reverse_lazy("developer_vendor_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        log_activity(self.request, "updated", "Developer/Vendor", f"Developer/vendor updated: {self.object.name}", self.object.pk)
        messages.success(self.request, "Developer/vendor updated.")
        return response


@login_required
def developer_vendor_delete(request, pk):
    vendor = get_object_or_404(DeveloperVendor.objects.annotate(assignment_count=Count("assignments", distinct=True)), pk=pk)
    linked_records = vendor.assignment_count
    if request.method == "POST":
        if linked_records:
            vendor.status = DeveloperVendor.VendorStatus.INACTIVE
            vendor.save(update_fields=["status", "updated_at"])
            log_activity(request, "deleted", "Developer/Vendor", f"Developer/vendor marked inactive: {vendor.name}", vendor.pk)
            messages.warning(
                request,
                "This developer/vendor is linked with existing project records. It cannot be permanently deleted. It was marked as Inactive.",
            )
        else:
            vendor_name = vendor.name
            vendor.delete()
            log_activity(request, "deleted", "Developer/Vendor", f"Developer/vendor deleted: {vendor_name}", pk)
            messages.success(request, f"Developer/vendor deleted: {vendor_name}")
        return redirect("developer_vendor_list")
    return render(
        request,
        "billing/confirm_delete.html",
        {
            "object_type": "developer/vendor",
            "object_name": vendor.name,
            "cancel_url": reverse_lazy("developer_vendor_list"),
            "linked_records": linked_records,
            "warning_message": (
                "This developer/vendor is linked with existing project records. It cannot be permanently deleted. "
                "Confirming will mark it as Inactive."
            )
            if linked_records
            else "",
            "confirm_label": "Mark Inactive" if linked_records else "Confirm Delete",
            "confirmation_text": (
                "Are you sure you want to delete this developer/vendor? "
                "This action cannot be undone if no records are linked."
            ),
        },
    )


@login_required
def month_summary(request):
    invoices = _final_invoice_queryset().select_related("company", "client").order_by("invoice_date")
    summaries = {}
    for invoice in invoices:
        key = (invoice.invoice_date.year, invoice.invoice_date.month, invoice.currency)
        if key not in summaries:
            first_day, last_day = _month_bounds(date(invoice.invoice_date.year, invoice.invoice_date.month, 1))
            summaries[key] = {
                "month": first_day,
                "last_day": last_day,
                "currency": invoice.currency,
                "total_invoices": 0,
                "raised": Decimal("0.00"),
                "received": Decimal("0.00"),
                "pending": Decimal("0.00"),
                "gst": Decimal("0.00"),
                "paid_count": 0,
                "pending_count": 0,
                "partial_count": 0,
            }
        summary = summaries[key]
        summary["total_invoices"] += 1
        summary["raised"] += invoice.total_amount
        summary["received"] += invoice.received_amount
        summary["pending"] += invoice.pending_amount
        summary["gst"] += invoice.gst_amount
        if invoice.payment_status == Invoice.PaymentStatus.PAID:
            summary["paid_count"] += 1
        elif invoice.payment_status == Invoice.PaymentStatus.PARTIALLY_PAID:
            summary["partial_count"] += 1
        else:
            summary["pending_count"] += 1
    rows = sorted(summaries.values(), key=lambda row: row["month"], reverse=True)
    return render(request, "billing/month_summary.html", {"rows": rows})


@login_required
def reports(request):
    context = _report_context(request.GET)
    return render(request, "billing/reports.html", context)


@login_required
def report_export_excel(request):
    try:
        output_path = _build_excel_report(_report_context(request.GET))
    except Exception as exc:
        logger.exception("Excel report generation failed")
        messages.error(request, f"Excel export failed: {exc}")
        return redirect("reports")
    return FileResponse(
        open(output_path, "rb"),
        as_attachment=True,
        filename=output_path.name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@login_required
def report_export_pdf(request):
    context = _report_context(request.GET)
    try:
        from weasyprint import HTML

        output_path = report_file_path(f"{_safe_report_base_name(context)}.pdf")
        html = render_to_string("reports/report_pdf.html", context)
        HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf(str(output_path))
    except Exception as exc:
        logger.exception("PDF report generation failed")
        messages.error(request, f"PDF report export failed: {exc}")
        return redirect("reports")
    return FileResponse(open(output_path, "rb"), as_attachment=True, filename=output_path.name, content_type="application/pdf")


@login_required
def project_reports(request):
    context = _project_report_context(request.GET)
    return render(request, "billing/project_reports.html", context)


@login_required
def project_report_export_excel(request):
    try:
        output_path = _build_project_excel_report(_project_report_context(request.GET))
    except Exception as exc:
        logger.exception("Project Excel report generation failed")
        messages.error(request, f"Project Excel export failed: {exc}")
        return redirect("project_reports")
    return FileResponse(
        open(output_path, "rb"),
        as_attachment=True,
        filename=output_path.name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@login_required
def project_report_export_pdf(request):
    context = _project_report_context(request.GET)
    try:
        from weasyprint import HTML

        output_path = report_file_path(f"{_safe_report_base_name(context)}.pdf")
        html = render_to_string("reports/project_report_pdf.html", context)
        HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf(str(output_path))
    except Exception as exc:
        logger.exception("Project PDF report generation failed")
        messages.error(request, f"Project PDF report export failed: {exc}")
        return redirect("project_reports")
    return FileResponse(open(output_path, "rb"), as_attachment=True, filename=output_path.name, content_type="application/pdf")


class RecurringInvoiceTemplateListView(PageSizeMixin, LoginRequiredMixin, ListView):
    model = RecurringInvoiceTemplate
    template_name = "billing/recurring_invoice_list.html"
    context_object_name = "templates_list"

    def get_queryset(self):
        return RecurringInvoiceTemplate.objects.select_related("company", "client", "project").prefetch_related("items")


@login_required
def recurring_invoice_create(request):
    template = RecurringInvoiceTemplate(next_invoice_date=timezone.localdate())
    if request.method == "POST":
        form = RecurringInvoiceTemplateForm(request.POST, instance=template)
        formset = RecurringInvoiceTemplateItemFormSet(request.POST, instance=template, prefix="items")
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                template = form.save()
                formset.instance = template
                formset.save()
            log_activity(request, "created", "Recurring Invoice", f"Recurring invoice template created: {template.title}", template.pk)
            messages.success(request, "Recurring invoice template saved.")
            return redirect("recurring_invoice_list")
    else:
        form = RecurringInvoiceTemplateForm(instance=template)
        formset = RecurringInvoiceTemplateItemFormSet(instance=template, prefix="items")
    return render(request, "billing/recurring_invoice_form.html", {"form": form, "formset": formset, "object": None})


@login_required
def recurring_invoice_edit(request, pk):
    template = get_object_or_404(RecurringInvoiceTemplate, pk=pk)
    if request.method == "POST":
        form = RecurringInvoiceTemplateForm(request.POST, instance=template)
        formset = RecurringInvoiceTemplateItemFormSet(request.POST, instance=template, prefix="items")
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                template = form.save()
                formset.save()
            log_activity(request, "updated", "Recurring Invoice", f"Recurring invoice template updated: {template.title}", template.pk)
            messages.success(request, "Recurring invoice template updated.")
            return redirect("recurring_invoice_list")
    else:
        form = RecurringInvoiceTemplateForm(instance=template)
        formset = RecurringInvoiceTemplateItemFormSet(instance=template, prefix="items")
    return render(request, "billing/recurring_invoice_form.html", {"form": form, "formset": formset, "object": template})


@login_required
@require_POST
def recurring_invoice_toggle(request, pk):
    template = get_object_or_404(RecurringInvoiceTemplate, pk=pk)
    template.is_active = not template.is_active
    template.save(update_fields=["is_active", "updated_at"])
    log_activity(
        request,
        "activated" if template.is_active else "deactivated",
        "Recurring Invoice",
        f"Recurring invoice template {'activated' if template.is_active else 'deactivated'}: {template.title}",
        template.pk,
    )
    messages.success(request, f"Recurring invoice template {'activated' if template.is_active else 'deactivated'}.")
    return redirect("recurring_invoice_list")


@login_required
@require_POST
def recurring_invoice_generate(request, pk):
    template = get_object_or_404(
        RecurringInvoiceTemplate.objects.select_related("company", "client", "project").prefetch_related("items__hsn_sac_code"),
        pk=pk,
    )
    item_rows = [
        {
            "description": item.description,
            "hsn_sac_code": item.hsn_sac_code,
            "item_price": item.item_price,
            "quantity": item.quantity,
        }
        for item in template.items.all()
    ]
    if not item_rows:
        messages.error(request, "Add at least one item before generating a draft invoice.")
        return redirect("recurring_invoice_edit", pk=template.pk)
    invoice_date = template.next_invoice_date or timezone.localdate()
    currency = template.project.currency if template.project_id else CURRENCY_INR
    calculated = calculate_invoice_totals(item_rows, template.company, apply_gst=template.apply_gst, currency=currency)
    app_settings = ApplicationSetting.load()
    invoice = Invoice(
        invoice_number=generate_draft_invoice_number(invoice_date),
        company=template.company,
        client=template.client,
        project=template.project,
        invoice_date=invoice_date,
        currency=currency,
        subject=template.title,
        apply_gst=bool(template.apply_gst and template.company.gstin),
        subtotal=calculated["subtotal"],
        gst_percentage=calculated["gst_percentage"],
        gst_amount=calculated["gst_amount"],
        total_amount=calculated["total_amount"],
        amount_in_words=calculated["amount_in_words"],
        terms_and_conditions=app_settings.default_terms_and_conditions,
        declaration=app_settings.default_declaration,
        payment_status=Invoice.PaymentStatus.PENDING,
        invoice_status=Invoice.InvoiceStatus.DRAFT,
        received_amount=to_money(0),
        pending_amount=calculated["total_amount"],
    )
    with transaction.atomic():
        invoice.save()
        InvoiceItem.objects.bulk_create(
            [
                InvoiceItem(
                    invoice=invoice,
                    serial_number=item["serial_number"],
                    description=item["description"],
                    hsn_sac_code=item.get("hsn_sac_code"),
                    item_price=item["item_price"],
                    quantity=item["quantity"],
                    total=item["total"],
                )
                for item in calculated["items"]
            ]
        )
        template.next_invoice_date = _advance_recurring_date(invoice_date, template.frequency)
        template.save(update_fields=["next_invoice_date", "updated_at"])
    log_activity(request, "draft generated", "Recurring Invoice", f"Draft invoice generated from template: {template.title}", invoice.pk)
    messages.success(request, f"Draft invoice generated: {invoice.invoice_number}")
    return redirect("invoice_edit", pk=invoice.pk)


@login_required
def settings_view(request):
    app_settings = ApplicationSetting.load()
    if request.method == "POST":
        form = ApplicationSettingsForm(request.POST, instance=app_settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Settings updated. New invoices will use the latest defaults.")
            return redirect("settings")
    else:
        form = ApplicationSettingsForm(instance=app_settings)

    return render(
        request,
        "billing/settings.html",
        {
            "form": form,
            "backup_location": "/backups/",
            "local_url": "http://127.0.0.1:8000",
            "logo_guidance": "Use PNG, JPG, JPEG, or SVG logos at least 300px by 150px and under 2 MB.",
        },
    )


@superuser_required
def backup_dashboard(request):
    return _render_backup_dashboard(request, BackupUploadForm())


@superuser_required
@require_POST
def backup_create(request):
    try:
        backup_path = create_local_backup()
    except BackupError as exc:
        logger.exception("Local backup creation failed")
        messages.error(request, f"Backup creation failed: {exc}")
    else:
        log_activity(request, "created", "Backup", f"Backup created: {backup_path.name}")
        messages.success(request, f"Backup created successfully: {backup_path.name}")
    return redirect("backup")


@superuser_required
def backup_download(request, filename):
    try:
        path = backup_file_path(filename)
    except BackupError as exc:
        messages.error(request, str(exc))
        return redirect("backup")
    return FileResponse(open(path, "rb"), as_attachment=True, filename=path.name, content_type="application/zip")


@superuser_required
@require_POST
def restore_upload(request):
    form = BackupUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Restore upload failed. Please choose a valid backup ZIP file.")
        return _render_backup_dashboard(request, form)

    try:
        token, validation = save_pending_restore(form.cleaned_data["backup_file"])
    except RestoreValidationError as exc:
        messages.error(request, str(exc))
        return _render_backup_dashboard(request, BackupUploadForm())

    request.session["pending_restore_token"] = token
    return render(
        request,
        "billing/restore_confirm.html",
        {
            "token": token,
            "validation": validation,
        },
    )


@superuser_required
def restore_confirm(request, token):
    if request.session.get("pending_restore_token") != token:
        messages.error(request, "Restore confirmation expired. Please upload the backup again.")
        return redirect("backup")

    try:
        path = pending_restore_path(token)
        validation = None
    except RestoreValidationError as exc:
        messages.error(request, str(exc))
        return redirect("backup")

    if request.method == "POST":
        if request.POST.get("confirm_restore") != "yes":
            messages.error(request, "Restore was not confirmed.")
            return redirect("backup")
        try:
            result = restore_local_backup(path)
        except (RestoreValidationError, RestoreError) as exc:
            logger.exception("Restore failed")
            messages.error(request, str(exc))
            return redirect("backup")
        cleanup_pending_restore(token)
        request.session.pop("pending_restore_token", None)
        messages.success(
            request,
            f"Restore completed. Safety backup created: {result.safety_backup_name}. Please log in again if prompted.",
        )
        log_activity(request, "completed", "Restore", f"Restore completed. Safety backup: {result.safety_backup_name}")
        return redirect("login")

    return render(request, "billing/restore_confirm.html", {"token": token, "validation": validation})


def _render_backup_dashboard(request, restore_form):
    return render(
        request,
        "billing/backup.html",
        {
            "restore_form": restore_form,
            "backups": list_local_backups(),
            "backup_location": "/backups/",
        },
    )


@login_required
@require_POST
def dismiss_backup_reminder(request):
    app_settings = ApplicationSetting.load()
    app_settings.backup_reminder_dismissed_on = timezone.localdate()
    app_settings.save(update_fields=["backup_reminder_dismissed_on", "updated_at"])
    messages.info(request, "Backup reminder dismissed for today.")
    return redirect(request.POST.get("next") or "dashboard")


def _backup_reminder_context(request):
    if not request.user.is_superuser:
        return None
    today = timezone.localdate()
    app_settings = ApplicationSetting.load()
    if app_settings.backup_reminder_dismissed_on == today:
        return None
    backups = list_local_backups()
    latest_backup = backups[0] if backups else None
    if latest_backup and latest_backup.created_at.date() > today - timedelta(days=7):
        return None
    return {
        "message": "Backup Reminder: It has been more than 7 days since your last backup. Please create a backup to protect your data.",
        "latest_backup": latest_backup,
    }


def _advance_recurring_date(current_date, frequency):
    months = {
        RecurringInvoiceTemplate.Frequency.MONTHLY: 1,
        RecurringInvoiceTemplate.Frequency.QUARTERLY: 3,
        RecurringInvoiceTemplate.Frequency.YEARLY: 12,
    }.get(frequency, 1)
    month_index = current_date.month - 1 + months
    year = current_date.year + month_index // 12
    month = month_index % 12 + 1
    day = min(current_date.day, calendar.monthrange(year, month)[1])
    return current_date.replace(year=year, month=month, day=day)


def _dashboard_project_queryset(data):
    projects = Project.objects.exclude(project_status=Project.ProjectStatus.CANCELLED)
    if data.get("client"):
        projects = projects.filter(client=data["client"])
    if data.get("currency"):
        projects = projects.filter(currency=data["currency"])
    return projects


def _invoice_summary_by_currency(invoices):
    return {currency: _invoice_summary(invoices.filter(currency=currency)) for currency, _label in CURRENCY_CHOICES}


def _project_summary_by_currency(projects, filters, invoices):
    summaries = {}
    for currency, _label in CURRENCY_CHOICES:
        currency_projects = projects.filter(currency=currency)
        project_summary = _project_summary(currency_projects)
        received = _dashboard_project_client_received_total(currency_projects, filters, invoices.filter(currency=currency))
        pending = to_money(project_summary["total_with_gst"] - received)
        if pending < Decimal("0.00"):
            pending = Decimal("0.00")
        summaries[currency] = {
            "total_with_gst": project_summary["total_with_gst"],
            "client_received": received,
            "client_pending": pending,
        }
    return summaries


def _dashboard_currency_summary_cards(invoice_currency_summaries, project_currency_summaries):
    cards = []
    for currency, _label in CURRENCY_CHOICES:
        invoice_summary = invoice_currency_summaries[currency]
        project_summary = project_currency_summaries[currency]
        invoice_received = invoice_summary["received"]
        invoice_pending = invoice_summary["pending"]
        total_received = to_money(invoice_received + project_summary["client_received"])
        cards.extend(
            [
                {"label": f"Raised invoice amount / Total Invoice Value - {currency}", "value": _money(invoice_summary["raised"], currency)},
                {"label": f"Invoice Received Amount - {currency}", "value": _money(invoice_received, currency)},
                {"label": f"Client / Project Received Amount - {currency}", "value": _money(project_summary["client_received"], currency)},
                {"label": f"Total Received Amount - {currency}", "value": _money(total_received, currency)},
                {"label": f"Invoice Pending Amount - {currency}", "value": _money(invoice_pending, currency)},
                {"label": f"Client / Project Pending Amount - {currency}", "value": _money(project_summary["client_pending"], currency)},
            ]
        )
    return cards


def _project_summary_card_list(summary):
    return [
        {"label": "Total projects", "value": str(summary["total_projects"])},
        {"label": "Draft projects", "value": str(summary["draft_projects"])},
        {"label": "Active / In Progress", "value": str(summary["active_projects"])},
        {"label": "Completed projects", "value": str(summary["completed_projects"])},
        {"label": "On Hold projects", "value": str(summary["on_hold_projects"])},
        {"label": "Recurring digital marketing", "value": str(summary["recurring_digital_marketing"])},
        {"label": "One-time digital marketing", "value": str(summary["one_time_digital_marketing"])},
        {"label": "Approved project value", "value": _money(summary["approved_value"])},
        {"label": "Client amount received", "value": _money(summary["client_received"])},
        {"label": "Client pending amount", "value": _money(summary["client_pending"])},
        {"label": "Project GST amount", "value": _money(summary["gst_amount"])},
        {"label": "Project total with GST", "value": _money(summary["total_with_gst"])},
        {"label": "Developer final cost", "value": _money(summary["developer_cost"])},
        {"label": "Developer amount paid", "value": _money(summary["developer_paid"])},
        {"label": "Developer pending amount", "value": _money(summary["developer_pending"])},
        {"label": "Estimated profit", "value": _money(summary["estimated_profit"])},
        {"label": "Actual cash profit", "value": _money(summary["actual_cash_profit"])},
    ]


def _dashboard_analytics_cards(invoice_summary, project_summary, invoice_currency_summaries, project_currency_summaries):
    cards = []
    for currency, _label in CURRENCY_CHOICES:
        currency_invoice = invoice_currency_summaries[currency]
        currency_project = project_currency_summaries[currency]
        cards.extend(
            [
                {"label": f"Total Invoice Value - {currency}", "value": _money(currency_invoice["raised"], currency), "class": "analytics-primary", "icon": currency},
                {"label": f"Invoice Received - {currency}", "value": _money(currency_invoice["received"], currency), "class": "analytics-success", "icon": currency},
                {"label": f"Invoice Pending - {currency}", "value": _money(currency_invoice["pending"], currency), "class": "analytics-warning", "icon": currency},
                {"label": f"Client / Project Received - {currency}", "value": _money(currency_project["client_received"], currency), "class": "analytics-success", "icon": currency},
            ]
        )
    return cards + [
        {"label": "Total Projects", "value": str(project_summary["total_projects"]), "class": "analytics-primary", "icon": "P"},
        {"label": "Active Projects", "value": str(project_summary["active_projects"]), "class": "analytics-info", "icon": "IP"},
        {"label": "Completed Projects", "value": str(project_summary["completed_projects"]), "class": "analytics-success", "icon": "C"},
        {"label": "On Hold", "value": str(project_summary["on_hold_projects"]), "class": "analytics-warning", "icon": "OH"},
        {"label": "Client Receivables", "value": _money(project_summary["client_pending"]), "class": "analytics-warning", "icon": "CR"},
        {"label": "Developer Payables", "value": _money(project_summary["developer_pending"]), "class": "analytics-danger", "icon": "DP"},
        {"label": "Actual Profit", "value": _money(project_summary["actual_cash_profit"]), "class": "analytics-success", "icon": "₹"},
        {"label": "Draft Invoices", "value": str(invoice_summary.get("draft_count", 0)), "class": "analytics-neutral", "icon": "DR"},
    ]


def _dashboard_chart_data(invoices, projects, project_summary, invoice_summary):
    status_labels = [choice[0] for choice in Project.ProjectStatus.choices]
    project_status_counts = [projects.filter(project_status=status).count() for status in status_labels]

    invoice_status_labels = [Invoice.PaymentStatus.PAID, Invoice.PaymentStatus.PENDING, Invoice.PaymentStatus.PARTIALLY_PAID, "Draft"]
    invoice_status_counts = [
        invoice_summary["paid_count"],
        invoice_summary["pending_count"],
        invoice_summary["partial_count"],
        invoice_summary.get("draft_count", 0),
    ]

    monthly = {}
    for invoice in _invoice_queryset_with_payment_totals(invoices).order_by("invoice_date"):
        key = invoice.invoice_date.replace(day=1)
        if key not in monthly:
            monthly[key] = {"raised": Decimal("0.00"), "received": Decimal("0.00"), "pending": Decimal("0.00")}
        received = invoice.payment_received or Decimal("0.00")
        pending = to_money(invoice.total_amount - received)
        if pending < Decimal("0.00"):
            pending = Decimal("0.00")
        monthly[key]["raised"] += invoice.total_amount
        monthly[key]["received"] += received
        monthly[key]["pending"] += pending

    month_keys = sorted(monthly.keys())[-6:]
    month_labels = [month.strftime("%b %Y") for month in month_keys]
    monthly_values = [monthly[month] for month in month_keys]

    return {
        "projectStatus": {
            "title": "Project Status",
            "labels": status_labels,
            "values": project_status_counts,
            "colors": ["#94a3b8", "#64748b", "#38bdf8", "#16a34a", "#2563eb", "#f59e0b", "#15803d", "#dc2626"],
        },
        "invoicePaymentStatus": {
            "title": "Invoice Payment Status",
            "labels": list(invoice_status_labels),
            "values": invoice_status_counts,
            "colors": ["#16a34a", "#dc2626", "#f59e0b", "#94a3b8"],
        },
        "fundStatus": {
            "title": "Client Receivable vs Developer Payable",
            "labels": ["Client Pending", "Developer Pending", "Client Received", "Developer Paid"],
            "values": [
                _chart_number(project_summary["client_pending"]),
                _chart_number(project_summary["developer_pending"]),
                _chart_number(project_summary["client_received"]),
                _chart_number(project_summary["developer_paid"]),
            ],
            "colors": ["#f59e0b", "#dc2626", "#16a34a", "#2563eb"],
        },
        "monthlyInvoice": {
            "title": "Monthly Invoice Raised vs Received",
            "labels": month_labels,
            "series": [
                {"label": "Raised", "values": [_chart_number(item["raised"]) for item in monthly_values], "color": "#2563eb"},
                {"label": "Received", "values": [_chart_number(item["received"]) for item in monthly_values], "color": "#16a34a"},
                {"label": "Pending", "values": [_chart_number(item["pending"]) for item in monthly_values], "color": "#f59e0b"},
            ],
        },
        "profitSummary": {
            "title": "Profit Summary",
            "labels": ["Approved Value", "Developer Cost", "Estimated Profit", "Actual Cash Profit"],
            "values": [
                _chart_number(project_summary["approved_value"]),
                _chart_number(project_summary["developer_cost"]),
                _chart_number(project_summary["estimated_profit"]),
                _chart_number(project_summary["actual_cash_profit"]),
            ],
            "colors": ["#2563eb", "#dc2626", "#7c3aed", "#16a34a"],
        },
    }


def _chart_number(value):
    return float(to_money(value or Decimal("0.00")))


def _project_linked_record_count(project):
    return (
        project.invoices.count()
        + project.client_payments.count()
        + project.assignments.count()
        + DeveloperPayment.objects.filter(project_assignment__project=project).count()
    )


def _project_summary(projects):
    project_list = list(projects.select_related("client").prefetch_related("assignments", "client_payments", "invoices"))
    totals = {
        "total_projects": len(project_list),
        "draft_projects": 0,
        "active_projects": 0,
        "completed_projects": 0,
        "on_hold_projects": 0,
        "recurring_digital_marketing": 0,
        "one_time_digital_marketing": 0,
        "approved_value": Decimal("0.00"),
        "client_received": Decimal("0.00"),
        "client_pending": Decimal("0.00"),
        "gst_amount": Decimal("0.00"),
        "total_with_gst": Decimal("0.00"),
        "developer_cost": Decimal("0.00"),
        "developer_paid": Decimal("0.00"),
        "developer_pending": Decimal("0.00"),
        "estimated_profit": Decimal("0.00"),
        "actual_cash_profit": Decimal("0.00"),
    }
    for project in project_list:
        if project.project_status == Project.ProjectStatus.DRAFT:
            totals["draft_projects"] += 1
        if project.project_type == Project.ProjectType.DIGITAL_MARKETING:
            if project.billing_type == Project.BillingType.RECURRING:
                totals["recurring_digital_marketing"] += 1
            else:
                totals["one_time_digital_marketing"] += 1
        if project.project_status in {Project.ProjectStatus.APPROVED, Project.ProjectStatus.IN_PROGRESS}:
            totals["active_projects"] += 1
        elif project.project_status == Project.ProjectStatus.COMPLETED:
            totals["completed_projects"] += 1
        elif project.project_status == Project.ProjectStatus.ON_HOLD:
            totals["on_hold_projects"] += 1
        summary = project_financial_summary(project)
        if project.project_status == Project.ProjectStatus.DRAFT:
            continue
        totals["approved_value"] += project.approved_quote or Decimal("0.00")
        client_received = to_money(sum((payment.amount_received or Decimal("0.00")) for payment in project.client_payments.all()))
        client_pending = to_money(summary["total_with_gst"] - client_received)
        if client_pending < Decimal("0.00"):
            client_pending = Decimal("0.00")
        totals["client_received"] += client_received
        totals["client_pending"] += client_pending
        totals["gst_amount"] += summary["gst_amount"]
        totals["total_with_gst"] += summary["total_with_gst"]
        totals["developer_cost"] += summary["developer_final_cost"]
        totals["developer_paid"] += summary["developer_paid"]
        totals["developer_pending"] += summary["developer_pending"]
        totals["estimated_profit"] += summary["estimated_profit"]
        totals["actual_cash_profit"] += to_money(client_received - summary["developer_paid"])
    return totals


def _project_report_context(params):
    data = params.copy()
    if not data:
        data["report_type"] = "project_wise"
    form = ProjectReportFilterForm(data)
    form.is_valid()
    cleaned = form.cleaned_data if form.is_valid() else {}
    report_type = cleaned.get("report_type") or "project_wise"
    projects = _apply_project_report_filters(
        Project.objects.select_related("client").prefetch_related(
            "client_payments",
            "assignments__developer_vendor",
            "assignments__developer_payments",
            "invoices",
        ),
        cleaned,
        report_type,
    )
    rows = project_report_rows(projects)
    context = {
        "form": form,
        "report_type": report_type,
        "report_title": _project_report_title(report_type),
        "generated_at": timezone.localtime(),
        "rows": rows,
        "summary": _project_rows_summary(rows),
        "currency_summaries": _project_rows_summary_by_currency(rows),
        "filter_details": _project_filter_details(cleaned),
    }
    return context


def _apply_project_report_filters(queryset, data, report_type):
    if data.get("start_date"):
        queryset = queryset.filter(start_date__gte=data["start_date"])
    if data.get("end_date"):
        queryset = queryset.filter(start_date__lte=data["end_date"])
    if data.get("client"):
        queryset = queryset.filter(client=data["client"])
    if data.get("project"):
        queryset = queryset.filter(pk=data["project"].pk)
    if data.get("project_type"):
        queryset = queryset.filter(project_type=data["project_type"])
    if data.get("project_status"):
        queryset = queryset.filter(project_status=data["project_status"])
    if data.get("currency"):
        queryset = queryset.filter(currency=data["currency"])
    if data.get("developer_vendor"):
        queryset = queryset.filter(assignments__developer_vendor=data["developer_vendor"])
    if data.get("min_completion") is not None:
        queryset = queryset.filter(completion_percentage__gte=data["min_completion"])
    if data.get("max_completion") is not None:
        queryset = queryset.filter(completion_percentage__lte=data["max_completion"])
    if data.get("payment_status"):
        queryset = queryset.filter(
            invoices__payment_status=data["payment_status"],
            invoices__invoice_status=Invoice.InvoiceStatus.FINAL,
            invoices__is_deleted=False,
        )
    if data.get("gst_type") == "gst":
        queryset = queryset.filter(invoices__apply_gst=True, invoices__gst_amount__gt=0, invoices__is_deleted=False)
    elif data.get("gst_type") == "non_gst":
        queryset = queryset.filter(
            Q(invoices__apply_gst=False) | Q(invoices__gst_amount=0),
            invoices__is_deleted=False,
        )
    if report_type == "pending_receivables":
        queryset = queryset.filter(client_pending_amount__gt=0)
    elif report_type == "pending_payables":
        queryset = queryset.filter(assignments__pending_amount_to_developer__gt=0)
    return queryset.distinct()


def _project_rows_summary(rows):
    summary = {
        "approved_value": Decimal("0.00"),
        "base_amount": Decimal("0.00"),
        "gst_amount": Decimal("0.00"),
        "total_with_gst": Decimal("0.00"),
        "client_received": Decimal("0.00"),
        "client_pending": Decimal("0.00"),
        "developer_cost": Decimal("0.00"),
        "developer_paid": Decimal("0.00"),
        "developer_pending": Decimal("0.00"),
        "invoice_raised": Decimal("0.00"),
        "invoice_received": Decimal("0.00"),
        "invoice_pending": Decimal("0.00"),
        "estimated_profit": Decimal("0.00"),
        "approved_profit": Decimal("0.00"),
        "actual_cash_profit": Decimal("0.00"),
    }
    for row in rows:
        project = row["project"]
        item = row["summary"]
        if project.project_status == Project.ProjectStatus.DRAFT:
            continue
        summary["approved_value"] += project.approved_quote or Decimal("0.00")
        summary["base_amount"] += item["base_amount"]
        summary["gst_amount"] += item["gst_amount"]
        summary["total_with_gst"] += item["total_with_gst"]
        summary["client_received"] += item["client_received"]
        summary["client_pending"] += item["client_pending"]
        summary["developer_cost"] += item["developer_final_cost"]
        summary["developer_paid"] += item["developer_paid"]
        summary["developer_pending"] += item["developer_pending"]
        summary["invoice_raised"] += item["invoice_raised"]
        summary["invoice_received"] += item["invoice_received"]
        summary["invoice_pending"] += item["invoice_pending"]
        summary["estimated_profit"] += item["estimated_profit"]
        summary["approved_profit"] += item["approved_profit"]
        summary["actual_cash_profit"] += item["actual_cash_profit"]
    return summary


def _project_rows_summary_by_currency(rows):
    return {
        currency: _project_rows_summary([row for row in rows if row["project"].currency == currency])
        for currency, _label in CURRENCY_CHOICES
    }


def _project_report_title(report_type):
    return dict(ProjectReportFilterForm.REPORT_TYPES).get(report_type, "Project report")


def _project_filter_details(data):
    details = []
    for label, key in [
        ("From", "start_date"),
        ("To", "end_date"),
        ("Client", "client"),
        ("Project", "project"),
        ("Project type", "project_type"),
        ("Project status", "project_status"),
        ("Currency", "currency"),
        ("Developer/vendor", "developer_vendor"),
        ("Min completion", "min_completion"),
        ("Max completion", "max_completion"),
        ("Invoice payment status", "payment_status"),
        ("GST Type", "gst_type"),
    ]:
        value = data.get(key)
        if value not in (None, ""):
            if key == "gst_type":
                value = _gst_type_label(value)
            details.append((label, value))
    return details


def _final_invoice_queryset():
    return Invoice.objects.filter(invoice_status=Invoice.InvoiceStatus.FINAL, is_deleted=False)


def _apply_gst_type_filter(queryset, gst_type):
    if gst_type == "gst":
        return queryset.filter(apply_gst=True, gst_amount__gt=0)
    if gst_type == "non_gst":
        return queryset.filter(Q(apply_gst=False) | Q(gst_amount=0))
    return queryset


def _apply_dashboard_filters(queryset, data):
    start_date, end_date = _dashboard_date_range(data)
    queryset = queryset.filter(invoice_date__gte=start_date, invoice_date__lte=end_date)
    if data.get("client"):
        queryset = queryset.filter(client=data["client"])
    if data.get("project"):
        queryset = queryset.filter(project=data["project"])
    if data.get("currency"):
        queryset = queryset.filter(currency=data["currency"])
    if data.get("payment_status"):
        queryset = queryset.filter(payment_status=data["payment_status"])
    return queryset


def _dashboard_received_total(invoices):
    return _dashboard_payments_queryset(invoices).aggregate(total=Sum("received_amount")).get("total") or Decimal("0.00")


def _dashboard_project_client_received_total(projects, data, invoices=None):
    start_date, end_date = _dashboard_date_range(data)
    active_projects = projects.exclude(
        project_status__in=[
            Project.ProjectStatus.DRAFT,
            Project.ProjectStatus.CANCELLED,
        ]
    )
    project_payments = ProjectClientPayment.objects.select_related("project").filter(
        project_id__in=active_projects.values("pk"),
        payment_date__gte=start_date,
        payment_date__lte=end_date,
    )
    if invoices is None:
        invoices = _final_invoice_queryset().filter(project_id__in=active_projects.values("pk"))
    invoice_payment_keys = Counter(
        (
            project_id,
            currency,
            payment_date,
            to_money(received_amount),
        )
        for project_id, currency, payment_date, received_amount in _dashboard_payments_queryset(invoices)
        .filter(invoice__project_id__isnull=False)
        .values_list("invoice__project_id", "invoice__currency", "payment_date", "received_amount")
    )
    total = Decimal("0.00")
    for project_id, currency, payment_date, amount_received in project_payments.values_list(
        "project_id",
        "project__currency",
        "payment_date",
        "amount_received",
    ):
        key = (project_id, currency, payment_date, to_money(amount_received))
        if invoice_payment_keys[key]:
            invoice_payment_keys[key] -= 1
            continue
        total += amount_received or Decimal("0.00")
    return to_money(total)


def _dashboard_payments_queryset(invoices):
    return Payment.objects.filter(
        invoice_id__in=invoices.values("pk"),
        invoice__invoice_status=Invoice.InvoiceStatus.FINAL,
        invoice__is_deleted=False,
    )


def _dashboard_date_range(data):
    today = timezone.localdate()
    period = data.get("period") or "this_month"
    if period == "last_month":
        first_this_month = today.replace(day=1)
        last_month_day = first_this_month - timedelta(days=1)
        return _month_bounds(last_month_day)
    if period == "custom" and data.get("start_date") and data.get("end_date"):
        return data["start_date"], data["end_date"]
    return _month_bounds(today)


def _month_bounds(day):
    first_day = day.replace(day=1)
    last_day = day.replace(day=calendar.monthrange(day.year, day.month)[1])
    return first_day, last_day


def _invoice_summary(queryset):
    aggregate = queryset.aggregate(
        raised=Sum("total_amount"),
        gst=Sum("gst_amount"),
        count=Count("id"),
    )
    raised = aggregate["raised"] or Decimal("0.00")
    received = _dashboard_received_total(queryset)
    pending = to_money(raised - received)
    if pending < Decimal("0.00"):
        pending = Decimal("0.00")
    status_counts = _invoice_payment_status_counts(queryset)
    return {
        "raised": raised,
        "received": received,
        "pending": pending,
        "gst": aggregate["gst"] or Decimal("0.00"),
        "count": aggregate["count"] or 0,
        "paid_count": status_counts["paid"],
        "pending_count": status_counts["pending"],
        "partial_count": status_counts["partial"],
    }


def _invoice_payment_status_counts(queryset):
    invoices = _invoice_queryset_with_payment_totals(queryset)
    paid = invoices.filter(total_amount__gt=0, payment_received__gte=F("total_amount")).count()
    partial = invoices.filter(
        total_amount__gt=0,
        payment_received__gt=0,
        payment_received__lt=F("total_amount"),
    ).count()
    pending = max(queryset.count() - paid - partial, 0)
    return {"paid": paid, "partial": partial, "pending": pending}


def _invoice_queryset_with_payment_totals(queryset):
    return queryset.annotate(
        payment_received=Coalesce(
            Sum("payments__received_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )


def _apply_invoice_filters(queryset, data):
    record_status = data.get("record_status") or "active"
    if record_status == "deleted":
        queryset = queryset.filter(is_deleted=True)
    elif record_status == "draft":
        queryset = queryset.filter(is_deleted=False, invoice_status=Invoice.InvoiceStatus.DRAFT)
    elif record_status == "final":
        queryset = queryset.filter(is_deleted=False, invoice_status=Invoice.InvoiceStatus.FINAL)
    elif record_status != "all":
        queryset = queryset.filter(is_deleted=False)

    query = data.get("q")
    if query:
        queryset = queryset.filter(
            Q(invoice_number__icontains=query)
            | Q(subject__icontains=query)
            | Q(company__company_name__icontains=query)
            | Q(client__client_name__icontains=query)
        )
    month_value = data.get("month")
    if month_value:
        try:
            year, month = [int(part) for part in month_value.split("-", 1)]
            queryset = queryset.filter(invoice_date__year=year, invoice_date__month=month)
        except (TypeError, ValueError):
            pass
    if data.get("start_date"):
        queryset = queryset.filter(invoice_date__gte=data["start_date"])
    if data.get("end_date"):
        queryset = queryset.filter(invoice_date__lte=data["end_date"])
    if data.get("company"):
        queryset = queryset.filter(company=data["company"])
    if data.get("client"):
        queryset = queryset.filter(client=data["client"])
    if data.get("project"):
        queryset = queryset.filter(project=data["project"])
    if data.get("currency"):
        queryset = queryset.filter(currency=data["currency"])
    if data.get("payment_status"):
        queryset = queryset.filter(payment_status=data["payment_status"])
    return _apply_gst_type_filter(queryset, data.get("gst_type"))


def _report_context(params):
    data = params.copy()
    if not data:
        data["report_type"] = "date_range"
    form = ReportFilterForm(data)
    form.is_valid()
    cleaned = form.cleaned_data if form.is_valid() else {}
    report_type = cleaned.get("report_type") or "date_range"
    invoice_base = Invoice.objects.select_related("company", "client").exclude(invoice_status=Invoice.InvoiceStatus.DRAFT)
    if not cleaned.get("include_deleted"):
        invoice_base = invoice_base.filter(invoice_status=Invoice.InvoiceStatus.FINAL, is_deleted=False)
    invoices = _apply_report_filters(invoice_base, cleaned, report_type)
    payment_base = Payment.objects.select_related("invoice", "invoice__company", "invoice__client").exclude(
        invoice__invoice_status=Invoice.InvoiceStatus.DRAFT,
    )
    if not cleaned.get("include_deleted"):
        payment_base = payment_base.filter(invoice__is_deleted=False)
    payments = _apply_payment_report_filters(payment_base, cleaned)
    if report_type != "payment_received":
        payments = Payment.objects.none()
    summary = _invoice_summary(invoices)
    if report_type == "payment_received":
        summary["received"] = payments.aggregate(total=Sum("received_amount")).get("total") or Decimal("0.00")
    currency_summaries = _report_currency_summaries(invoices, payments, report_type)
    context = {
        "form": form,
        "report_type": report_type,
        "report_title": _report_title(report_type),
        "generated_at": timezone.localtime(),
        "invoices": invoices,
        "payments": payments,
        "summary": summary,
        "currency_summaries": currency_summaries,
        "filter_details": _filter_details(cleaned),
    }
    return context


def _report_currency_summaries(invoices, payments, report_type):
    summaries = {}
    for currency, _label in CURRENCY_CHOICES:
        currency_invoices = invoices.filter(currency=currency)
        summary = _invoice_summary(currency_invoices)
        if report_type == "payment_received":
            summary["received"] = payments.filter(invoice__currency=currency).aggregate(total=Sum("received_amount")).get("total") or Decimal("0.00")
        summaries[currency] = summary
    return summaries


def _apply_report_filters(queryset, data, report_type):
    if data.get("start_date"):
        queryset = queryset.filter(invoice_date__gte=data["start_date"])
    if data.get("end_date"):
        queryset = queryset.filter(invoice_date__lte=data["end_date"])
    if data.get("company"):
        queryset = queryset.filter(company=data["company"])
    if data.get("client"):
        queryset = queryset.filter(client=data["client"])
    if data.get("currency"):
        queryset = queryset.filter(currency=data["currency"])
    if data.get("payment_status"):
        queryset = queryset.filter(payment_status=data["payment_status"])
    if report_type == "paid":
        queryset = queryset.filter(payment_status=Invoice.PaymentStatus.PAID)
    elif report_type == "pending":
        queryset = queryset.filter(payment_status=Invoice.PaymentStatus.PENDING)
    elif report_type == "partially_paid":
        queryset = queryset.filter(payment_status=Invoice.PaymentStatus.PARTIALLY_PAID)
    elif report_type == "gst":
        queryset = queryset.filter(apply_gst=True, gst_amount__gt=0)
    return _apply_gst_type_filter(queryset, data.get("gst_type"))


def _apply_payment_report_filters(queryset, data):
    if data.get("start_date"):
        queryset = queryset.filter(payment_date__gte=data["start_date"])
    if data.get("end_date"):
        queryset = queryset.filter(payment_date__lte=data["end_date"])
    if data.get("company"):
        queryset = queryset.filter(invoice__company=data["company"])
    if data.get("client"):
        queryset = queryset.filter(invoice__client=data["client"])
    if data.get("currency"):
        queryset = queryset.filter(invoice__currency=data["currency"])
    if data.get("payment_status"):
        queryset = queryset.filter(invoice__payment_status=data["payment_status"])
    if data.get("gst_type") == "gst":
        queryset = queryset.filter(invoice__apply_gst=True, invoice__gst_amount__gt=0)
    elif data.get("gst_type") == "non_gst":
        queryset = queryset.filter(Q(invoice__apply_gst=False) | Q(invoice__gst_amount=0))
    return queryset


def _filter_details(data):
    details = []
    for label, key in [
        ("From", "start_date"),
        ("To", "end_date"),
        ("Company", "company"),
        ("Client", "client"),
        ("Currency", "currency"),
        ("Payment status", "payment_status"),
        ("GST Type", "gst_type"),
    ]:
        value = data.get(key)
        if value:
            if key == "gst_type":
                value = _gst_type_label(value)
            details.append((label, value))
    if data.get("include_deleted"):
        details.append(("Deleted invoices", "Included"))
    return details


def _gst_type_label(value):
    return {
        "gst": "GST Invoices",
        "non_gst": "Non-GST Invoices",
    }.get(value, "All")


def _report_title(report_type):
    return dict(ReportFilterForm.REPORT_TYPES).get(report_type, "Invoice report")


def _safe_report_base_name(context):
    today = timezone.localdate().strftime("%d%m%Y")
    title = "".join(ch if ch.isalnum() else "_" for ch in context["report_title"]).strip("_")
    return f"{title}_{today}"


def _build_excel_report(context):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    title_fill = PatternFill("solid", fgColor="E9F2EF")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    worksheet["A1"] = context["report_title"]
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = f"Generated: {context['generated_at'].strftime('%d-%m-%Y %H:%M')}"
    row = 4
    worksheet.cell(row=row, column=1, value="Filters").font = bold
    for label, value in context["filter_details"]:
        row += 1
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=str(value))
    row += 2

    if context["report_type"] == "payment_received":
        headers = ["Payment Date", "Invoice Number", "Company", "Client", "Currency", "Received Amount", "Payment Mode", "Remarks"]
        rows = [
            [
                payment.payment_date,
                payment.invoice.invoice_number,
                payment.invoice.company.company_name,
                payment.invoice.client.client_name,
                payment.invoice.currency,
                format_currency(payment.received_amount, payment.invoice.currency),
                payment.payment_mode,
                payment.remarks,
            ]
            for payment in context["payments"]
        ]
    else:
        headers = [
            "Invoice Number",
            "Invoice Date",
            "Company",
            "Client",
            "Subject",
            "Currency",
            "Subtotal",
            "GST Amount",
            "GST Type",
            "Total Amount",
            "Received Amount",
            "Pending Amount",
            "Payment Status",
        ]
        rows = [
            [
                invoice.invoice_number,
                invoice.invoice_date,
                invoice.company.company_name,
                invoice.client.client_name,
                invoice.subject,
                invoice.currency,
                format_currency(invoice.subtotal, invoice.currency),
                format_currency(invoice.gst_amount, invoice.currency),
                "GST" if invoice.apply_gst and invoice.gst_amount > 0 else "Non-GST",
                format_currency(invoice.total_amount, invoice.currency),
                format_currency(invoice.received_amount, invoice.currency),
                format_currency(invoice.pending_amount, invoice.currency),
                invoice.payment_status,
            ]
            for invoice in context["invoices"]
        ]

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=column, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_values in rows:
        row += 1
        for column, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row, column=column, value=value)
            if isinstance(value, Decimal):
                cell.number_format = '#,##0.00'
    row += 2
    totals = []
    for currency, summary in context["currency_summaries"].items():
        totals.extend(
            [
                (f"Total raised amount - {currency}", format_currency(summary["raised"], currency)),
                (f"Total received amount - {currency}", format_currency(summary["received"], currency)),
                (f"Total pending amount - {currency}", format_currency(summary["pending"], currency)),
                (f"Total GST amount - {currency}", format_currency(summary["gst"], currency)),
            ]
        )
    for label, value in totals:
        worksheet.cell(row=row, column=1, value=label).font = bold
        worksheet.cell(row=row, column=2, value=value)
        worksheet.cell(row=row, column=1).fill = title_fill
        worksheet.cell(row=row, column=2).fill = title_fill
        row += 1

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)

    output_path = report_file_path(f"{_safe_report_base_name(context)}.xlsx")
    workbook.save(output_path)
    return output_path


def _build_project_excel_report(context):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Project Report"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="E9F2EF")
    bold = Font(bold=True)

    worksheet["A1"] = context["report_title"]
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = f"Generated: {context['generated_at'].strftime('%d-%m-%Y %H:%M')}"
    row = 4
    worksheet.cell(row=row, column=1, value="Filters").font = bold
    for label, value in context["filter_details"]:
        row += 1
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=str(value))
    row += 2

    headers = [
        "Project ID",
        "Project Name",
        "Client",
        "Project Type",
        "Currency",
        "Approved Quote",
        "GST Type",
        "Base Amount",
        "GST %",
        "GST Applicable Amount",
        "GST Amount",
        "Total With GST",
        "Client Received",
        "Client Pending",
        "Developer/Vendor",
        "Developer Final Cost",
        "Developer Paid",
        "Developer Pending",
        "Invoice Raised",
        "Invoice Received",
        "Invoice Pending",
        "Status",
        "Completion %",
        "Estimated Profit",
        "Approved Profit",
        "Actual Cash Profit",
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=column, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for item in context["rows"]:
        project = item["project"]
        summary = item["summary"]
        row += 1
        values = [
            project.project_id,
            project.project_name,
            project.client.client_name,
            project.custom_project_type or project.project_type,
            project.currency,
            format_currency(project.approved_quote, project.currency),
            summary["gst_type_label"],
            format_currency(summary["base_amount"], project.currency),
            summary["gst_percentage"],
            format_currency(summary["partial_gst_taxable_amount"], project.currency),
            format_currency(summary["gst_amount"], project.currency),
            format_currency(summary["total_with_gst"], project.currency),
            format_currency(summary["client_received"], project.currency),
            format_currency(summary["client_pending"], project.currency),
            summary["developer_names"],
            format_currency(summary["developer_final_cost"], project.currency),
            format_currency(summary["developer_paid"], project.currency),
            format_currency(summary["developer_pending"], project.currency),
            format_currency(summary["invoice_raised"], project.currency),
            format_currency(summary["invoice_received"], project.currency),
            format_currency(summary["invoice_pending"], project.currency),
            project.project_status,
            project.completion_percentage,
            format_currency(summary["estimated_profit"], project.currency),
            format_currency(summary["approved_profit"], project.currency),
            format_currency(summary["actual_cash_profit"], project.currency),
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column, value=value)
            if isinstance(value, Decimal):
                cell.number_format = '#,##0.00'

    row += 2
    totals = []
    for currency, summary in context["currency_summaries"].items():
        totals.extend(
            [
                (f"Approved project value - {currency}", format_currency(summary["approved_value"], currency)),
                (f"Total with GST - {currency}", format_currency(summary["total_with_gst"], currency)),
                (f"Client amount received - {currency}", format_currency(summary["client_received"], currency)),
                (f"Client pending amount - {currency}", format_currency(summary["client_pending"], currency)),
                (f"Actual cash profit - {currency}", format_currency(summary["actual_cash_profit"], currency)),
            ]
        )
    for label, value in totals:
        worksheet.cell(row=row, column=1, value=label).font = bold
        worksheet.cell(row=row, column=2, value=value)
        worksheet.cell(row=row, column=1).fill = title_fill
        worksheet.cell(row=row, column=2).fill = title_fill
        row += 1

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 44)

    output_path = report_file_path(f"{_safe_report_base_name(context)}.xlsx")
    workbook.save(output_path)
    return output_path
